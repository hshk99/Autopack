"""
Excel Export Service - Generate organized document pack as Excel workbook
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from datetime import datetime
from sqlalchemy.orm import Session
from app.models.document import Document
from app.models.category import Category
from typing import List


class ExcelExportService:
    def export_pack(
        self,
        pack_name: str,
        documents: List[Document],
        categories: List[Category],
        output_path: Path
    ) -> Path:
        """
        Export documents as organized Excel workbook
        """
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create Summary sheet
        summary_sheet = wb.create_sheet("Summary")
        summary_sheet['A1'] = pack_name
        summary_sheet['A1'].font = Font(size=16, bold=True)

        summary_sheet['A3'] = "Generated:"
        summary_sheet['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M')

        summary_sheet['A4'] = "Total Documents:"
        summary_sheet['B4'] = len(documents)

        summary_sheet['A5'] = "Categories:"
        summary_sheet['B5'] = len(categories)

        summary_sheet['A6'] = "Approved Documents:"
        summary_sheet['B6'] = len([d for d in documents if d.classification_confidence == 100])

        # Format summary sheet
        for row in summary_sheet['A3:A6']:
            row[0].font = Font(bold=True)

        # Create All Documents sheet
        all_docs_sheet = wb.create_sheet("All Documents")

        # Headers
        headers = ['File Name', 'Category', 'Confidence', 'Status', 'File Size (KB)']
        for col_num, header in enumerate(headers, 1):
            cell = all_docs_sheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        category_map = {cat.id: cat.name for cat in categories}

        for row_num, doc in enumerate(documents, 2):
            all_docs_sheet.cell(row=row_num, column=1, value=doc.filename)

            category_name = category_map.get(doc.assigned_category_id, 'Uncategorized')
            all_docs_sheet.cell(row=row_num, column=2, value=category_name)

            confidence = f"{doc.classification_confidence:.0f}%" if doc.classification_confidence else 'N/A'
            all_docs_sheet.cell(row=row_num, column=3, value=confidence)

            status = 'Approved' if doc.classification_confidence == 100 else 'Pending'
            all_docs_sheet.cell(row=row_num, column=4, value=status)

            file_size_kb = doc.file_size / 1024
            all_docs_sheet.cell(row=row_num, column=5, value=f"{file_size_kb:.2f}")

        # Auto-adjust column widths
        for column in all_docs_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            all_docs_sheet.column_dimensions[column_letter].width = adjusted_width

        # Create sheets for each category
        category_docs = {}
        for cat in categories:
            category_docs[cat.id] = {
                'name': cat.name,
                'documents': []
            }

        for doc in documents:
            if doc.assigned_category_id and doc.assigned_category_id in category_docs:
                category_docs[doc.assigned_category_id]['documents'].append(doc)

        for cat_id, cat_data in category_docs.items():
            if not cat_data['documents']:
                continue

            # Create category sheet
            sheet_name = cat_data['name'][:31]  # Excel sheet name limit
            cat_sheet = wb.create_sheet(sheet_name)

            # Headers
            for col_num, header in enumerate(['File Name', 'Confidence', 'Status'], 1):
                cell = cat_sheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            # Data
            for row_num, doc in enumerate(cat_data['documents'], 2):
                cat_sheet.cell(row=row_num, column=1, value=doc.filename)
                confidence = f"{doc.classification_confidence:.0f}%" if doc.classification_confidence else 'N/A'
                cat_sheet.cell(row=row_num, column=2, value=confidence)
                status = 'Approved' if doc.classification_confidence == 100 else 'Pending'
                cat_sheet.cell(row=row_num, column=3, value=status)

            # Auto-adjust widths
            for column in cat_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                cat_sheet.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(output_path)
        return output_path
